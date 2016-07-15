#changes to implement
import openpyxl
from openpyxl.styles import PatternFill
#spreadsheet 1 is the old spreadsheet
#spreadsheet 2 is the spreadsheet you want to identify new values in

def col2num(col_str):
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char.upper()) - ord('A') + 1) * (26 ** expn)
        expn += 1
    return col_num

directory1 = input("Enter in the file path of the first excel sheet (Case sensitive): \n")
sheetname1 = input("Enter in the name of the sheet: \n")
column1 = col2num(input("Enter the column letter of the data: \n"))
directory2 = input("Enter in the file path of the second excel sheet: \n")
sheetname2 = input("Enter in the name of the sheet: \n")
column2 = col2num(input("Enter the column letter of the data: \n"))
wb1 = openpyxl.load_workbook(directory1)
sheet1 = wb1.get_sheet_by_name(sheetname1)
wb2 = openpyxl.load_workbook(directory2)
sheet2 = wb2.get_sheet_by_name(sheetname2)

rgb=[255,255,0]
color_string="".join([str(hex(x))[2:].upper().rjust(2, "0") for x in rgb])

i=2 #start from the second row in the second sheet
while (sheet2.cell(row=i, column = column1).value!= None) & (sheet2.cell(row=i+1, column = column1).value!= None):
    i=i+1
    #value = sheet2.cell(row=275, column=2).value
    j=2
    match = 0 #if match is equal to zero there are no matches. Set match equal to one if there is a match
    while (sheet1.cell(row=j, column = column2).value != None):
        j=j+1
        if (sheet2.cell(row=i, column = column2).value == sheet1.cell(row=j, column = column1).value):
            #print(str(i) + "," +str(j)+"\n")
            match = 1
    if match == 0:
        #print(sheet2.cell(row=i, column=1).value)
        sheet2.cell(row=i, column=column1).fill=PatternFill(fill_type="solid", start_color='FF' + color_string, end_color='FF' + color_string)
folderpath = '\\'.join(directory2.split('\\')[0:-1])
#print(folderpath)
filepath = folderpath + "\\out.xlsx"
#print(filepath)
wb2.save(filepath)
